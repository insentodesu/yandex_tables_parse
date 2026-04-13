"""Build an XLSX template for importing the source structure into Yandex Tables."""

from __future__ import annotations

import csv
import io
import json
import ssl
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import config
from table_client import normalize_header

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


def _download_text(url: str) -> str:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 accounting-max-bot"},
    )
    context = ssl.create_default_context()
    if not config.MAX_SSL_VERIFY:
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(request, timeout=30, context=context) as response:
        return response.read().decode("utf-8-sig")


def _fetch_headers() -> list[str]:
    content = _download_text(config.SOURCE_GOOGLE_CSV_URL)
    reader = csv.reader(io.StringIO(content))
    first_row = next(reader)
    return [normalize_header(value) for value in first_row]


def _write_structure_snapshot(headers: list[str]) -> None:
    payload = {
        "source_csv_url": config.SOURCE_GOOGLE_CSV_URL,
        "column_count": len(headers),
        "headers": headers,
        "command_column": config.TABLE_COMMAND_COLUMN,
        "command_options": config.CHAT_OPTIONS,
    }
    output_path = Path(config.SOURCE_STRUCTURE_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _build_workbook(headers: list[str]) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = config.TABLE_SHEET_NAME or "Рабочая таблица"
    worksheet.freeze_panes = "A2"
    options_sheet = workbook.create_sheet(title="_options")
    options_sheet.sheet_state = "hidden"

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        worksheet.column_dimensions[cell.column_letter].width = max(14, min(len(header) + 4, 40))

    if config.TABLE_COMMAND_COLUMN not in headers:
        raise ValueError(f"Column not found: {config.TABLE_COMMAND_COLUMN}")

    command_column_index = headers.index(config.TABLE_COMMAND_COLUMN) + 1
    for row_index, option in enumerate(config.CHAT_OPTIONS, start=1):
        options_sheet.cell(row=row_index, column=1, value=option)

    validation = DataValidation(
        type="list",
        formula1=f"=_options!$A$1:$A${len(config.CHAT_OPTIONS)}",
        allow_blank=True,
        showDropDown=True,
    )
    validation.prompt = "Выберите формат уведомления для бухгалтера"
    worksheet.add_data_validation(validation)
    validation.add(f"{worksheet.cell(row=1, column=command_column_index).column_letter}2:"
                   f"{worksheet.cell(row=1, column=command_column_index).column_letter}10000")
    return workbook


def main() -> None:
    headers = _fetch_headers()
    _write_structure_snapshot(headers)
    workbook = _build_workbook(headers)
    output_path = Path(config.TEMPLATE_OUTPUT_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
