"""Helpers for loading rows from exported spreadsheet sources."""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
from collections import Counter
import random
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from python_calamine import load_workbook as load_calamine_workbook

import config

logger = logging.getLogger(__name__)

_RETRIABLE_HTTP_CODES = frozenset({429, 502, 503})
_YANDEX_DISK_OAUTH_TYPES = frozenset({"yandex_disk_xlsx", "yandex_disk_csv"})


def normalize_header(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\n", " ").replace("\r", " ")
    return " ".join(text.split()).strip()


def normalize_cell(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(text.split()).strip()


def _format_http_error_response(code: int, body: bytes) -> str:
    msg: str
    try:
        data = json.loads(body.decode("utf-8"))
        if isinstance(data, dict):
            err = str(data.get("error") or "").strip()
            m = str(data.get("message") or data.get("description") or "").strip()
            parts = [p for p in (err, m) if p]
            if parts:
                msg = f"Yandex Disk API HTTP {code}: {' — '.join(parts)}"
            else:
                compact = json.dumps(data, ensure_ascii=False)[:400]
                msg = f"Yandex Disk API HTTP {code}: {compact}"
        else:
            raise ValueError("not a dict")
    except Exception:
        snippet = body.decode("utf-8", errors="replace").strip().replace("\n", " ")[:500]
        if code == 429:
            tail = "too many requests (rate limited)"
        elif code == 403:
            tail = "forbidden"
        else:
            tail = "request failed"
        msg = f"Yandex Disk API HTTP {code}: {tail}" + (f" — {snippet}" if snippet else "")

    if code == 403:
        msg += (
            " | Проверьте пароль (TABLE_YANDEX_PUBLIC_PASSWORD), путь внутри папки "
            "(TABLE_YANDEX_PUBLIC_PATH), что ссылка живая; при «без скачивания» у публичной ссылки "
            "перейдите на TABLE_SOURCE_TYPE=yandex_disk_xlsx и YANDEX_DISK_TOKEN + TABLE_DISK_PATH."
        )
    return msg


def _merge_url_query(url: str, updates: dict[str, str]) -> str:
    parsed = urllib.parse.urlparse(url)
    pairs = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
    merged = dict(pairs)
    merged.update(updates)
    new_query = urllib.parse.urlencode(merged)
    return urllib.parse.urlunparse(parsed._replace(query=new_query))


def _retry_after_seconds(exc: urllib.error.HTTPError) -> float | None:
    if exc.headers is None:
        return None
    raw = exc.headers.get("Retry-After")
    if raw is None or not str(raw).strip():
        return None
    raw = str(raw).strip()
    try:
        return float(raw)
    except ValueError:
        pass
    try:
        when = parsedate_to_datetime(raw)
        if when.tzinfo is None:
            when = when.replace(tzinfo=timezone.utc)
        return max(0.0, (when - datetime.now(timezone.utc)).total_seconds())
    except Exception:
        return None


@dataclass(slots=True)
class SpreadsheetRow:
    sheet_name: str
    row_number: int
    values: dict[str, str]


class TableClient:
    """Reads CSV/XLSX from a URL, file, public Yandex link, or OAuth path on Yandex Disk."""

    def __init__(
        self,
        source_type: str | None = None,
        source: str | None = None,
        sheet_name: str | None = None,
        command_column: str | None = None,
    ) -> None:
        self.source_type = (source_type or config.TABLE_SOURCE_TYPE).strip().lower()
        self.source = (source or config.TABLE_SOURCE).strip()
        self.sheet_name = (sheet_name or config.TABLE_SHEET_NAME).strip()
        self.command_column = normalize_header(command_column or config.TABLE_COMMAND_COLUMN)

    async def get_rows(self) -> list[SpreadsheetRow]:
        return await asyncio.to_thread(self._get_rows_sync)

    async def get_actionable_rows(self) -> list[SpreadsheetRow]:
        rows = await self.get_rows()
        return [
            row for row in rows
            if normalize_cell(row.values.get(self.command_column, ""))
        ]

    def _get_rows_sync(self) -> list[SpreadsheetRow]:
        if self.source_type not in _YANDEX_DISK_OAUTH_TYPES and not self.source:
            raise ValueError("TABLE_SOURCE is not configured")

        if self.source_type == "csv_url":
            return self._load_csv(self._download_bytes(self.source).decode("utf-8-sig"))
        if self.source_type == "csv_file":
            return self._load_csv(Path(self.source).read_text(encoding="utf-8-sig"))
        if self.source_type == "xlsx_file":
            return self._load_xlsx(Path(self.source).read_bytes())
        if self.source_type == "xlsx_url":
            return self._load_xlsx(self._download_bytes(self.source))
        if self.source_type == "yandex_public_xlsx":
            return self._load_xlsx(self._download_yandex_public_bytes(self.source))
        if self.source_type == "yandex_public_csv":
            return self._load_csv(
                self._download_yandex_public_bytes(self.source).decode("utf-8-sig")
            )
        if self.source_type == "yandex_disk_xlsx":
            return self._load_xlsx(self._download_yandex_disk_oauth_bytes())
        if self.source_type == "yandex_disk_csv":
            return self._load_csv(self._download_yandex_disk_oauth_bytes().decode("utf-8-sig"))
        raise ValueError(f"Unsupported TABLE_SOURCE_TYPE: {self.source_type}")

    def _request_headers(self, extra: dict[str, str] | None = None) -> dict[str, str]:
        headers = {
            "User-Agent": config.HTTP_USER_AGENT,
            "Accept": "*/*",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
            # Снижаем риск отдачи старого XLSX с CDN/прокси при частом опросе.
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }
        if extra:
            headers.update(extra)
        return headers

    def _download_bytes(self, url: str, extra_headers: dict[str, str] | None = None) -> bytes:
        request = urllib.request.Request(url, headers=self._request_headers(extra_headers))
        context = ssl.create_default_context()
        if not config.MAX_SSL_VERIFY:
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE

        max_attempts = config.TABLE_FETCH_MAX_RETRIES
        base = config.TABLE_FETCH_RETRY_BASE_SECONDS
        cap = config.TABLE_FETCH_RETRY_MAX_SLEEP_SECONDS

        for attempt in range(max_attempts):
            try:
                with urllib.request.urlopen(request, timeout=30, context=context) as response:
                    return response.read()
            except urllib.error.HTTPError as exc:
                body = exc.read()
                if exc.code in _RETRIABLE_HTTP_CODES and attempt < max_attempts - 1:
                    hint = _retry_after_seconds(exc)
                    if hint is None:
                        hint = base * (2**attempt) + random.random()
                    wait = max(1.0, min(hint, cap))
                    logger.warning(
                        "HTTP %s при загрузке таблицы, повтор через %.1f с (попытка %s/%s)",
                        exc.code,
                        wait,
                        attempt + 1,
                        max_attempts,
                    )
                    time.sleep(wait)
                    continue
                raise RuntimeError(_format_http_error_response(exc.code, body)) from exc

    def _yandex_public_download_params_variants(self, public_key: str) -> list[dict[str, str]]:
        base: dict[str, str] = {"public_key": public_key}
        inner_path = config.TABLE_YANDEX_PUBLIC_PATH.strip()
        if inner_path:
            base["path"] = inner_path
        password = config.TABLE_YANDEX_PUBLIC_PASSWORD.strip()
        if not password:
            return [base]
        fixed_key = config.TABLE_YANDEX_PUBLIC_PASSWORD_PARAM.strip()
        if fixed_key:
            merged = dict(base)
            merged[fixed_key] = password
            return [merged]
        variants: list[dict[str, str]] = []
        for key in ("password", "pass", "pwd", "link_password"):
            merged = dict(base)
            merged[key] = password
            variants.append(merged)
        return variants

    def _download_yandex_public_bytes(self, public_key: str) -> bytes:
        variants = self._yandex_public_download_params_variants(public_key)
        payload: dict[str, Any] | None = None
        last_api_err: BaseException | None = None
        for params in variants:
            api_url = (
                "https://cloud-api.yandex.net/v1/disk/public/resources/download?"
                + urllib.parse.urlencode(params)
            )
            try:
                payload = self._fetch_json(api_url)
                break
            except RuntimeError as exc:
                last_api_err = exc
                text = str(exc)
                if "403" not in text and "401" not in text:
                    raise
                logger.warning(
                    "Публичный API отклонил запрос (возможен неверный ключ пароля в query); пробуем следующий вариант"
                )
        if payload is None:
            if last_api_err is not None:
                raise RuntimeError(
                    f"{last_api_err} "
                    "Если у ссылки отключено скачивание, публичный API Яндекса часто отвечает 403 даже с паролем — "
                    "включите скачивание в настройках ссылки или перейдите на "
                    "TABLE_SOURCE_TYPE=yandex_disk_xlsx с YANDEX_DISK_TOKEN и TABLE_DISK_PATH."
                ) from last_api_err
            raise RuntimeError("Yandex public download: no API response")

        download_url = normalize_cell(payload.get("href", ""))
        if not download_url:
            raise ValueError("Yandex public resource download URL is missing")
        cdn_headers: dict[str, str] = {}
        pk = public_key.strip()
        if pk.startswith("http://") or pk.startswith("https://"):
            cdn_headers["Referer"] = pk
        hdr = cdn_headers if cdn_headers else None
        pwd = config.TABLE_YANDEX_PUBLIC_PASSWORD.strip()

        try:
            return self._download_bytes(download_url, hdr)
        except RuntimeError as exc:
            if not pwd or "403" not in str(exc):
                raise
            logger.warning("CDN вернул 403; пробуем добавить пароль в query ссылки скачивания")
            for qkey in ("password", "pass", "key", "code"):
                try:
                    merged_url = _merge_url_query(download_url, {qkey: pwd})
                    return self._download_bytes(merged_url, hdr)
                except RuntimeError as exc2:
                    if "403" not in str(exc2):
                        raise
            raise exc

    def _download_yandex_disk_oauth_bytes(self) -> bytes:
        token = config.YANDEX_DISK_TOKEN.strip()
        disk_path = config.TABLE_DISK_PATH.strip()
        if not token:
            raise ValueError("YANDEX_DISK_TOKEN is not configured")
        if not disk_path:
            raise ValueError("TABLE_DISK_PATH is not configured")
        auth_headers = {"Authorization": f"OAuth {token}"}
        api_url = (
            "https://cloud-api.yandex.net/v1/disk/resources/download?"
            + urllib.parse.urlencode({"path": disk_path})
        )
        payload = self._fetch_json(api_url, auth_headers)
        download_url = normalize_cell(payload.get("href", ""))
        if not download_url:
            raise ValueError("Yandex Disk OAuth download URL is missing")
        return self._download_bytes(download_url, auth_headers)

    def _fetch_json(self, url: str, extra_headers: dict[str, str] | None = None) -> dict[str, Any]:
        merged = dict(extra_headers or {})
        if "cloud-api.yandex.net" in url:
            merged["Accept"] = "application/json"
        raw = self._download_bytes(url, merged if merged else None)
        data = json.loads(raw.decode("utf-8"))
        if not isinstance(data, dict):
            raise ValueError("Expected JSON object from Yandex public API")
        return data

    def _load_csv(self, content: str) -> list[SpreadsheetRow]:
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return []

        headers = [normalize_header(header) for header in rows[0]]
        result: list[SpreadsheetRow] = []
        for row_number, row in enumerate(rows[1:], start=2):
            if not any(normalize_cell(cell) for cell in row):
                continue
            padded = list(row) + [""] * max(0, len(headers) - len(row))
            values = {
                headers[idx]: normalize_cell(padded[idx])
                for idx in range(len(headers))
                if headers[idx]
            }
            result.append(
                SpreadsheetRow(
                    sheet_name=self.sheet_name or "Sheet1",
                    row_number=row_number,
                    values=values,
                )
            )
        return result

    def _load_xlsx(self, content: bytes) -> list[SpreadsheetRow]:
        if not content:
            raise ValueError("XLSX: пустой ответ при скачивании (0 байт)")

        openpyxl_errors: list[str] = []

        for read_only in (True, False):
            try:
                bio = io.BytesIO(content)
                workbook = load_workbook(
                    filename=bio,
                    data_only=True,
                    read_only=read_only,
                )
                try:
                    return self._load_xlsx_rows_with_openpyxl(workbook)
                finally:
                    closer = getattr(workbook, "close", None)
                    if callable(closer):
                        try:
                            closer()
                        except Exception:
                            pass
            except Exception as exc:
                openpyxl_errors.append(f"read_only={read_only}: {exc!r}")
                logger.debug(
                    "openpyxl не смог прочитать XLSX (read_only=%s): %s",
                    read_only,
                    exc,
                )

        try:
            rows = self._load_xlsx_rows_with_calamine(content)
            if openpyxl_errors:
                logger.info(
                    "XLSX прочитан через calamine (openpyxl часто падает на файлах из Яндекса: invalid XML/stylesheet — это не ошибка, если строки загрузились)"
                )
            return rows
        except Exception as exc_cal:
            joined = "; ".join(openpyxl_errors)
            logger.error(
                "calamine не смог прочитать XLSX: %s | openpyxl до этого: %s",
                exc_cal,
                joined,
            )
            raise RuntimeError(
                "Не удалось прочитать XLSX (ни openpyxl, ни calamine). "
                f"Openpyxl: {joined}. Calamine: {exc_cal!r}"
            ) from exc_cal

    def _load_xlsx_rows_with_openpyxl(self, workbook: Any) -> list[SpreadsheetRow]:
        result: list[SpreadsheetRow] = []
        for worksheet in self._select_openpyxl_worksheets(workbook):
            rows = list(worksheet.iter_rows(values_only=True))
            result.extend(self._build_spreadsheet_rows(worksheet.title, rows))
        return result

    def _load_xlsx_rows_with_calamine(self, content: bytes) -> list[SpreadsheetRow]:
        workbook = load_calamine_workbook(io.BytesIO(content))
        result: list[SpreadsheetRow] = []
        for worksheet_name in self._select_sheet_names(workbook.sheet_names):
            worksheet = workbook.get_sheet_by_name(worksheet_name)
            result.extend(self._build_spreadsheet_rows(worksheet_name, worksheet.to_python()))
        return result

    def _select_openpyxl_worksheets(self, workbook: Any) -> list[Any]:
        selected_names = self._select_sheet_names(tuple(workbook.sheetnames))
        if selected_names:
            return [workbook[name] for name in selected_names]
        return [workbook.active]

    def _select_sheet_names(self, sheet_names: tuple[str, ...] | list[str]) -> list[str]:
        if self.sheet_name:
            return [self.sheet_name]
        month_sheets = [name for name in config.MONTH_SHEET_NAMES if name in sheet_names]
        if month_sheets:
            return month_sheets
        if sheet_names:
            return [sheet_names[0]]
        return []

    def _build_spreadsheet_rows(
        self,
        sheet_name: str,
        rows: list[list[Any]] | list[tuple[Any, ...]],
    ) -> list[SpreadsheetRow]:
        if not rows:
            return []

        headers = [normalize_header(cell) for cell in rows[0]]
        dup_count = Counter(h for h in headers if h)
        for h, c in dup_count.items():
            if c > 1:
                logger.warning(
                    "Дублируется заголовок столбца после нормализации: %r (%s раз). "
                    "Значения в dict ячеек перезаписываются — колонка «%s» может читаться не из той ячейки.",
                    h,
                    c,
                    config.TABLE_COMMAND_COLUMN,
                )
        result: list[SpreadsheetRow] = []
        for row_number, row in enumerate(rows[1:], start=2):
            if not any(normalize_cell(cell) for cell in row):
                continue
            padded = list(row) + [""] * max(0, len(headers) - len(row))
            values = {
                headers[idx]: normalize_cell(padded[idx])
                for idx in range(len(headers))
                if headers[idx]
            }
            result.append(
                SpreadsheetRow(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    values=values,
                )
            )
        return result
